import pandas as pd
import os
import openpyxl

tables = {
    "san_pham": [
        [1, "id", "Long/BIGINT", "PK, Auto Increment", "ID sản phẩm"],
        [2, "ten_sp", "String/VARCHAR(255)", "Not Null", "Tên sản phẩm"],
        [3, "mo_ta", "TEXT", "", "Mô tả sản phẩm"],
        [4, "gia_ban", "BigDecimal/DECIMAL", "", "Giá bán"],
        [5, "noi_bat", "Boolean/BIT", "Default false", "Trạng thái nổi bật"],
        [6, "loai_sp_id", "Long/BIGINT", "FK (loai_sp)", "Khóa ngoại liên kết loại sản phẩm"]
    ],
    "gio_hang": [
        [1, "id", "Long/BIGINT", "PK, Auto Increment", "ID giỏ hàng"],
        [2, "tong_tien_tam_tinh", "BigDecimal/DECIMAL", "", "Tổng tiền tạm tính"],
        [3, "tk_id", "Long/BIGINT", "FK (tai_khoan), Unique", "Khóa ngoại liên kết tài khoản"]
    ],
    "tai_khoan": [
        [1, "id", "Long/BIGINT", "PK, Auto Increment", "ID tài khoản"],
        [2, "ten_dang_nhap", "String/VARCHAR", "Not Null, Unique", "Tên đăng nhập"],
        [3, "mat_khau", "String/VARCHAR", "Not Null", "Mật khẩu"]
    ],
    "gio_hang_ct": [
        [1, "gio_hang_id", "Long/BIGINT", "PK, FK (gio_hang)", "Khóa ngoại giỏ hàng"],
        [2, "sp_ct_id", "Long/BIGINT", "PK, FK (san_pham_ct)", "Khóa ngoại sản phẩm chi tiết"],
        [3, "so_luong_du_dinh", "Integer/INT", "", "Số lượng dự định mua"],
        [4, "thanh_tien", "BigDecimal/DECIMAL", "", "Thành tiền"]
    ],
    "hinh_anh_sp_ct": [
        [1, "id", "Long/BIGINT", "PK, Auto Increment", "ID hình ảnh"],
        [2, "link_hinh_anh", "String/VARCHAR", "Unique with sp_ct_id", "Link hình ảnh"],
        [3, "sp_ct_id", "Long/BIGINT", "FK (san_pham_ct), Unique with link_hinh_anh", "Khóa ngoại sản phẩm chi tiết"]
    ],
    "loai_sp": [
        [1, "id", "Long/BIGINT", "PK, Auto Increment", "ID loại sản phẩm"],
        [2, "ten_loai", "String/VARCHAR", "Not Null", "Tên loại"],
        [3, "slug", "String/VARCHAR", "Unique", "Slug URL"]
    ],
    "mau_sac": [
        [1, "id", "String/VARCHAR", "PK", "Mã màu (vd: #000000)"],
        [2, "ten_mau", "String/VARCHAR", "", "Tên màu"]
    ],
    "khuyen_mai": [
        [1, "ma_giam_gia", "String/VARCHAR", "PK", "Mã giảm giá"],
        [2, "ten_khuyen_mai", "String/VARCHAR", "Not Null", "Tên khuyến mãi"],
        [3, "gia_tri_giam", "BigDecimal/DECIMAL", "Not Null", "Giá trị giảm"],
        [4, "ngay_bat_dau", "LocalDateTime/DATETIME", "", "Ngày bắt đầu"],
        [5, "ngay_ket_thuc", "LocalDateTime/DATETIME", "", "Ngày kết thúc"],
        [6, "trang_thai", "Integer/INT", "", "Trạng thái (1: Active, 0: Inactive)"],
        [7, "don_toi_thieu", "BigDecimal/DECIMAL", "", "Đơn tối thiểu để áp dụng"],
        [8, "so_luot_sd", "Integer/INT", "", "Số lượt sử dụng"]
    ],
    "su_dung_khuyen_mai": [
        [1, "ma_su_dung", "Long/BIGINT", "PK, Auto Increment", "Mã sử dụng KM"],
        [2, "ma_tk", "Long/BIGINT", "FK (tai_khoan)", "Khóa ngoại tài khoản"],
        [3, "ma_don_mua", "Long/BIGINT", "FK (don_mua)", "Khóa ngoại đơn mua"],
        [4, "ma_giam_gia", "String/VARCHAR", "FK (khuyen_mai)", "Khóa ngoại khuyến mãi"],
        [5, "ngay_su_dung", "LocalDateTime/DATETIME", "", "Ngày sử dụng"],
        [6, "so_luot_sd", "Integer/INT", "", "Số lượt đã dùng"]
    ],
    "thong_tin_giao_hang": [
        [1, "id", "Long/BIGINT", "PK, Auto Increment", "ID thông tin giao hàng"],
        [2, "ten_kh", "String/VARCHAR", "", "Tên khách hàng"],
        [3, "sdt_kh", "String/VARCHAR", "", "Số điện thoại KH"],
        [4, "dia_chi", "String/VARCHAR", "", "Địa chỉ giao hàng"],
        [5, "mac_dinh", "Boolean/BIT", "", "Đặt làm mặc định"],
        [6, "trang_thai", "String/VARCHAR", "Default 'Đang dùng'", "Trạng thái"],
        [7, "tk_id", "Long/BIGINT", "FK (tai_khoan), Not Null", "Khóa ngoại tài khoản"]
    ],
    "san_pham_ct": [
        [1, "id", "Long/BIGINT", "PK, Auto Increment", "ID SP chi tiết"],
        [2, "sp_id", "Long/BIGINT", "FK (san_pham), Unique with mau_id", "Khóa ngoại sản phẩm"],
        [3, "mau_id", "String/VARCHAR", "FK (mau_sac), Unique with sp_id", "Khóa ngoại màu sắc"],
        [4, "so_luong_ton", "Integer/INT", "", "Số lượng tồn kho"],
        [5, "sku", "String/VARCHAR", "", "Mã SKU"]
    ],
    "don_mua_ct": [
        [1, "don_mua_id", "Long/BIGINT", "PK, FK (don_mua)", "Khóa ngoại đơn mua"],
        [2, "sp_ct_id", "Long/BIGINT", "PK, FK (san_pham_ct)", "Khóa ngoại sản phẩm chi tiết"],
        [3, "so_luong_mua", "Integer/INT", "", "Số lượng mua"],
        [4, "thanh_tien", "BigDecimal/DECIMAL", "", "Thành tiền"]
    ],
    "don_mua": [
        [1, "id", "Long/BIGINT", "PK, Auto Increment", "ID đơn mua"],
        [2, "ngay_tao_don", "LocalDateTime/DATETIME", "", "Ngày tạo đơn"],
        [3, "ma_giam_gia", "String/VARCHAR", "FK (khuyen_mai)", "Khóa ngoại khuyến mãi"],
        [4, "ty_le_chiet_khau", "Double/FLOAT", "", "Tỷ lệ chiết khấu"],
        [5, "tong_tien", "BigDecimal/DECIMAL", "", "Tổng tiền"],
        [6, "tong_thanh_toan", "BigDecimal/DECIMAL", "", "Tổng thanh toán"],
        [7, "trang_thai", "String/VARCHAR", "", "Trạng thái đơn hàng"],
        [8, "ly_do_huy", "String/VARCHAR", "", "Lý do hủy đơn"],
        [9, "ngay_huy", "LocalDateTime/DATETIME", "", "Ngày hủy đơn"],
        [10, "ghi_chu", "String/VARCHAR", "", "Ghi chú"],
        [11, "ttgh_id", "Long/BIGINT", "FK (thong_tin_giao_hang)", "Khóa ngoại thông tin giao hàng"]
    ]
}

columns = ["STT", "Thuộc tính", "Loại dữ liệu", "Ràng buộc", "Ghi chú"]

with pd.ExcelWriter("Database_Schema.xlsx", engine='openpyxl') as writer:
    row_offset = 0
    workbook = writer.book
    worksheet = workbook.create_sheet("Schema")
    
    for table_name, data in tables.items():
        # Write Table Name
        worksheet.cell(row=row_offset + 1, column=1, value=table_name.upper())
        worksheet.cell(row=row_offset + 1, column=1).font = openpyxl.styles.Font(bold=True, size=14)
        
        # Write Headers
        for col_idx, col_name in enumerate(columns):
            cell = worksheet.cell(row=row_offset + 2, column=col_idx + 1, value=col_name)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'), 
                right=openpyxl.styles.Side(style='thin'), 
                top=openpyxl.styles.Side(style='thin'), 
                bottom=openpyxl.styles.Side(style='thin')
            )
            
        # Write Data
        for row_idx, row_data in enumerate(data):
            for col_idx, val in enumerate(row_data):
                cell = worksheet.cell(row=row_offset + 3 + row_idx, column=col_idx + 1, value=val)
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'), 
                    right=openpyxl.styles.Side(style='thin'), 
                    top=openpyxl.styles.Side(style='thin'), 
                    bottom=openpyxl.styles.Side(style='thin')
                )
        
        row_offset += len(data) + 3

    # Adjust column widths
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

    # Remove default sheet
    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

print("Excel file created successfully.")
